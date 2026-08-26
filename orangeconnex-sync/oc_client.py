"""OrangeConnex client — login via Playwright, fetch data via in-page API.

Login through the real UI to establish a proper session, then navigate
to a business page and call the backend APIs directly from within that
page's JS context (avoids "system busy" errors from bare requests).
"""

import json
from io import BytesIO
import time

from openpyxl import load_workbook

from playwright.sync_api import sync_playwright

from config import OC_PASSWORD, OC_USERNAME

EF = "https://ef-web-sg.orangeconnex.com"
BASE = "https://fulfillment-cn.orangeconnex.com"

INVENTORY_PATH = "/api/app/seller/weaver-service/sku/warehouse/v1/inventory/query"
STATEMENT_PATH = "/api/app/seller/weaver-service/sku/warehouse/v1/statement/query"
AGING_EXPORT_PATH = "/api/app/seller-bff/export/v1/excel/async"
DOWNLOAD_CENTER_PATH = "/api/app/seller-bff/downloadCenter/v1/query"
INBOUND_PATH = "/api/app/seller/slark-inbound/inbound/v1/queryInboundOrderList"
ORDER_PATH = "/api/app/seller-bff/meepo/outboundOrder/v1/order/page"
REGION_PATH = (
    "/api/app/seller/pudge-service/seller/mfservice/v2/"
    "seller/joined/mfService/wareshoueRegion"
)


class OCClient:
    def __init__(self):
        self._pw = sync_playwright().start()
        self._browser = self._pw.chromium.launch(headless=True)
        self._ctx = self._browser.new_context(locale="zh-CN")
        self.page = self._ctx.new_page()
        self._login()

    # ---------- session ----------
    def _login(self):
        page = self.page
        page.goto(f"{BASE}/seller/login", timeout=60000)
        page.wait_for_load_state("domcontentloaded")
        time.sleep(3)
        dialog = page.locator(".el-dialog__wrapper:visible button:has-text('确认选择')")
        if dialog.count() > 0:
            dialog.first.click()
            time.sleep(1)
        page.fill("input[type='text']", OC_USERNAME)
        page.fill("input[type='password']", OC_PASSWORD)
        page.click("button:has-text('登录')")
        time.sleep(8)
        # Navigate to an authenticated page so API calls work
        page.goto(f"{BASE}/order", timeout=60000)
        time.sleep(5)

    def _token(self) -> str:
        match = self.page.evaluate(
            "document.cookie.match(/mf-seller-token=([^;]+)/)"
        )
        return match[1] if match else ""

    def _fetch(self, path: str, body=None) -> dict:
        token = self._token()
        body_json = json.dumps(body) if body else None
        method = "POST" if body is not None else "GET"
        result = self.page.evaluate(
            """async ([url, body, method, token]) => {
                const opts = {
                    method,
                    headers: { token },
                };
                if (body) {
                    opts.headers['Content-Type'] = 'application/json';
                    opts.body = body;
                }
                const r = await fetch(url, opts);
                return await r.json();
            }""",
            [EF + path, body_json, method, token],
        )
        if not isinstance(result, dict):
            raise RuntimeError(f"OC API {path} returned non-dict")
        if not result.get("success"):
            raise RuntimeError(
                f"OC API {path} failed: {json.dumps(result, ensure_ascii=False)[:300]}"
            )
        return result

    def list_warehouse_regions(self):
        data = self._fetch(REGION_PATH)
        regions = data.get("result") or []
        return [
            r["warehouseRegion"]
            for r in regions
            if r.get("hasWarehouseAllocation") and r.get("warehouseRegion")
        ]

    def _paginate(self, path, base_body, result_key="list", limit=5000):
        """Paginated fetch with large page size; returns all rows."""
        rows = []
        page_num = 1
        while True:
            body = dict(base_body, page=page_num, limit=limit)
            result = self._fetch(path, body)
            payload = result.get("result") or {}
            batch = payload.get(result_key) or []
            rows.extend(batch)
            total = payload.get("total") or 0
            if len(rows) >= total or not batch:
                break
            page_num += 1
            time.sleep(0.3)
        return rows

    # ---------- public ----------
    def list_inventory(self, region: str):
        body = {
            "warehouseRegionCode": region, "skuId": "", "skuIdList": [],
            "skuName": "", "sortFiled": "", "dimensionTypeCode": "",
            "sortRules": "", "inventorySymbol": "", "inventoryValue": None,
            "salesDays": 30,
        }
        return self._paginate(INVENTORY_PATH, body, result_key="list")

    def list_statement(self, region: str, from_ms: int, to_ms: int):
        body = {
            "warehouseRegionCode": region, "skuIdList": [], "tradeType": "",
            "createdTimestampFrom": from_ms, "createdTimestampTo": to_ms,
        }
        return self._paginate(STATEMENT_PATH, body, result_key="list")

    def list_inventory_aging(self, timeout_seconds: int = 600):
        """Export SKU aging for all warehouses and parse the generated Excel."""
        user_code = self.page.evaluate(
            "() => JSON.parse(localStorage.userProfile || '{}').code || ''"
        )
        if not user_code:
            raise RuntimeError("OC user code is not available")

        condition = {
            "skuId": "", "skuIdList": [], "skuName": "", "sortFiled": "",
            "dimensionTypeCode": "", "sortRules": "", "inventorySymbol": "",
            "inventoryValue": None, "salesDays": 30, "userCode": user_code,
            "language": "zh-CN",
            "warehouseRegionCodes": self.list_warehouse_regions(),
        }
        self._fetch(
            AGING_EXPORT_PATH,
            {
                "exportType": "inventory-aging-list-export",
                "exportCondition": json.dumps(
                    condition, separators=(",", ":"), ensure_ascii=False
                ),
            },
        )

        deadline = time.time() + timeout_seconds
        title = ""
        while time.time() < deadline:
            time.sleep(10)
            downloads = self._fetch(DOWNLOAD_CENTER_PATH, {"page": 1, "limit": 10})
            tasks = (downloads.get("result") or {}).get("list") or []
            if tasks and not title:
                title = tasks[0].get("title")
            task = next((item for item in tasks if item.get("title") == title), None)
            if not task or not task.get("downloadPath"):
                continue

            response = self._ctx.request.get(task["downloadPath"])
            if not response.ok:
                continue
            workbook = load_workbook(BytesIO(response.body()), data_only=True)
            worksheet = workbook.active
            excel_rows = list(worksheet.iter_rows(values_only=True))
            if len(excel_rows) < 2:
                continue

            headers = list(excel_rows[0])
            return [
                dict(zip(headers, row))
                for row in excel_rows[1:]
                if any(value is not None and value != "" for value in row)
            ]
        raise TimeoutError(f"OC aging export did not finish within {timeout_seconds}s")

    def list_inbound(self):
        body = {
            "inboundOrderStatus": "-1", "orderNoType": "inboundNum",
            "inboundNum": "", "trackingNum": "",
            "inboundReferenceNum": None, "inboundOrderNumbers": None,
            "trackingNumbers": None, "inboundReferenceNumbers": "",
            "warehouseArea": "", "warehouseCode": "", "sellerSKUId": "",
            "inboundDocumentAuditStatus": "", "vatNo": "", "asnId": "",
            "timeType": "createTime", "createTime": "",
            "realityArrivalDate": "", "listingTime": "", "pickupTime": "",
            "inboundMethod": "", "inboundDocumentIsRequired": False,
            "haveExceptionIsRequired": False, "orderSkuQtyDiff": None,
            "createTimeStartDate": "", "createTimeEndDate": "",
            "arrivalTimeStartDate": "", "arrivalTimeEndDate": "",
            "putAwayStartDate": "", "putAwayEndDate": "",
            "pickupStartDate": "", "pickupEndDate": "",
        }
        return self._paginate(INBOUND_PATH, body, result_key="data")

    def list_orders(self, days: int = 30):
        now_ms = int(time.time() * 1000)
        from_ms = now_ms - days * 86400 * 1000
        body = {
            "orderNoType": "orderNumber", "orderNumber": "",
            "trackingNo": "", "transactionId": "", "platformOrderId": "",
            "skuType": "skuId", "skuId": "", "sellerSkuId": "",
            "ebayOrderNumber": "", "referenceNumber": "",
            "submissionDate": [from_ms, now_ms],
            "submissionDateStart": from_ms,
            "submissionDateEnd": now_ms,
        }
        return self._paginate(ORDER_PATH, body, result_key="list", limit=1000)

    def close(self):
        self._browser.close()
        self._pw.stop()

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        self.close()
